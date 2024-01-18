import pandas as pd
import numpy
import json
import glob
import datetime
import openpyxl
import watchdog
import traceback
import shutil
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.utils.dataframe import dataframe_to_rows


now = datetime.datetime.now()
timestamp = now.strftime("%Y%m%d_%H%M%S")





try:

    excel_files = [ file for file in glob.glob('C:/Users/ASUS/Desktop/fieldmobi/Upload Folder/*.xlsx') if 'Error_Template' not in file]
    print(f'Excel files found: {excel_files}')
    if excel_files:

        for file in excel_files:
            df = pd.read_excel(file, engine = 'openpyxl', nrows = 3)
            print(df)


    def script1():
        if excel_files:
            for file in excel_files:
                df = pd.read_excel(file, engine = 'openpyxl', nrows = 2)
                data = df.iloc[0,[2,4]]
                data_str = str(data.iloc[0]) + '_' + str(data.iloc[1])
                print(data_str)

                df.pd.read_excel(file, engine = 'openpyxl', skiprows = 3)
                df = df.drop(df.columns[0], axis = 1)
                df = df.dropna(how = 'all')

                df = df.rename(column = {
                    'Field Name': 'data',
                    'Default Label': 'label',
                    'List Type': 'list_type',
                    'Default List Value': 'list_value'
                
                })

                if df['data'].isnull().any():
                    raise ValueError("Misisng value in 'data' column")
                df = df.drop(columns = ['Field Type'])
                print(df)


                json_obj = {}
                for i, row in df.iterrows():
                    row_data = {}
                    for col in df.columns:
                        if pd.notna(row[col]):
                            row_data[col] = row[col]

                    json_obj['fieldCode' + str(i+1).zfill(3)] = row_data
                json_str = json.dumps(json_obj, indent = 4)

                with open('output_fieldlist.txt', 'w') as file:
                    file.write(data_str + '\n\n')
                    file.write(json_str)
                print(json_str)
        else:
            print(" ")



    def script2():
        if excel_files:
            for file in excel_files:
                df = pd.read_excel(file, engine = 'openpyxl', nrows = 3)
                print(df)
                data1 = df.iloc[0,2]
                data2 = df.iloc[1,2]
                data3 = df.iloc[1,4]
                data_str = str(data1) + '_' + str(data2) + '_' + str(data3)
                print(data_str)


                df = pd.read_excel(file, engine = 'openpyxl', skiprows = 3)
                df = df.drop(df.columns[0], axis=1)
                df = df.dropna(how = 'all')
                df.reset_index(drop = True, inplace = True)

                df = df.rename(columns={
                    'Field Name' : 'data',
                    'New Label': 'label'

                })

                if df['data'].isnull().any():
                    raise ValueError("Missing value in data:")
                
                df.rename(columns = {
                    'Mobile Seq': 'Mobile_Mobile Seq',
                    'Validation': 'Mobile_Validation',
                    'Link Setup': 'Mobile_Link Setup',
                    'Update Setup': 'Mobile_Update Setup'
                }, inplace = True)
                
                df.rename(columns = {
                    'Search Seq': 'Web_Search Seq',
                    'Display Seq': 'Web_Display Seq',
                    'Create Seq': 'Web_Create Seq',
                    'Edit Seq': 'Web_Edit Seq',
                    'Validation.1': 'Web_Validation',
                    'Link Setup.1': 'Web_Link Setup',
                    'Update Setup.1': 'Web_Update Setup',
                    'List Seq': 'Web_List Seq',
                    'Summary Seq': 'Web_Summary Seq',
                    'Map Seq': 'Web_Map Seq',
                    'Report Seq': 'Web_Report Seq'
                }, inplace = True)
                
                json_obj = {}
                for i, row in df.iterrows():
                    row_data = {}
                    for col in df.columns:
                        if pd.notna(row[col]):
                            row_data[col] = row[col]
                    json_obj['fieldCode' + str(i+1).zfill(3)] = row_data
                json_str = json.dumps(json_obj, indent = 4)

                with open('output_dataview.txt', 'w') as file:
                    file.write(data_str + '\n\n')
                    file.write(json_str)
                print(json_str)

            else:
                print("No Excel files found in the local directory")


    def script3():
        if excel_files:
            for  file in excel_files:
                df = pd.read_excel(file, engine = 'openpyxl', nrows = 3)
                data1 = df.iloc[0,2]
                data2 = df.iloc[1,2]
                data3 = df.iloc[1,4]
                data_str = str(data1) + '_' + str(data2) + '_' + str(data3)
                print(data_str)


                df = pd.read_excel(file, engine = 'openpyxl', skiprows = 5)
                df= df.drop(df.columns[0], axis = 1)
                df =df.dropna(how = 'all')
                print(df)


                if df['KEY'].isnull().any() or df['TYP'].isnull().any():
                    print("Missing value in 'KEY' or 'TYP' column.")

                validation_row = pd.read_excel(file, engine='openpyxl', skiprows=5, nrows=1, usecols =lambda x: x not in ['Unnamed: 0', 'Unnamed: 1'])
                mandatory_columns = validation_row.columns[validation_row.eq('mandatory').any()]

                print(validation_row)
                print("Mandatory Columns:", mandatory_columns)

                df.reset_index(drop = True, inplace = True)
                missing_values = df[mandatory_columns].isnull()
                df_1 = df[~missing_values.any(axis=1)]




                json_obj = {}
                for i, row in df_1.iloc[1:].iterrows():
                    row_data = {}
                    for col in df.columns:
                        if pd.notna(row[col]):
                            row_data[col] = row[col]
                        json_obj[row['KEY']] = row_data
                json_str = json.dumps(json_obj, indent = 4)

                with open('output_datatemplate.txt', 'w') as file:
                    file.write(data_str + '\n\n')
                    file.write(json_str)

                missing_values = df[mandatory_columns].isnull()
                df_errors = df[missing_values.any(axis = 1)]
                wb = load_workbook('Error_template.xlsx')
                print(wb.sheetnames)
                ws = wb.active
                print(ws)
                df  =pd.read_excel('Error_template.xlsx', sheet_name = ws.title)
                print(df)

                current_row = 8
                for index, row in df_errors.iterrows():
                    for i, value in enumerate(row):
                        cell = ws.cell(row=current_row, column = i+2)
                        if pd.isnull(value):
                            if ws.cell(row = 6, column = i+2).value in ['KEY', 'TYP']:
                                cell.font = Font(color = "FF0000")
                                cell.value = "critical Error"
                            else:
                                cell.value = "Mandatory Data"
                        else:
                            cell.value = value
                    current_row += 1
                wb.save(f'C:\\Users\\ASUS\\Desktop\\fieldmobi\\Error Folder\\Error_{timestamp}.xlsx')
                

        else:
            print("No Excel files found in the local directory")
    
    
        if excel_files:
            for file in excel_files:
                df = pd.read_excel(file, engine = 'openpyxl', nrows = 5)  # Increase nrows to 5 to read the first 5 rows

                print(df.iloc[0,1], df.iloc[0,3])
                if df.iloc[0,1] == 'Application' and df.iloc[0,3] == 'Module':
                    script1()
                elif df.iloc[1,7] == 'Mobile Configuration' and df.iloc[1,11] == 'Web Configuration':
                    script2()
                elif df.iloc[4,2] == 'KEY' and df.iloc[4,3] == 'TYP':
                    script3()


except Exception as e:
    print(f"An error occured: {e}")  
    traceback.print_exc()


