import pandas as pd
import json
import glob
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.utils.dataframe import dataframe_to_rows

now = datetime.datetime.now()
timestamp = now.strftime("%Y%m%d_%H%M%S")




try:
    excel_files = [ file for file in glob.glob('C:/Users/ASUS/Desktop/fieldmobi/Upload Folder/*.xlsx') if 'Error_Template' not in file]
    print(f'Excel files found: {excel_files}')

    if excel_files:
        for file in excel_files:
            df = pd.read_excel(file, engine='openpyxl', nrows = 3)
            print(df)
            data1 = df.iloc[0,2]
            data2 = df.iloc[1,2]
            data3 = df.iloc[1,4]
            data_str = str(data1) + '_' + str(data2) + '_' + str(data3)
            print(data_str)

            df = pd.read_excel(file, engine='openpyxl', skiprows=5)
            df = df.drop(df.columns[0], axis=1)
            df = df.dropna(how = 'all')
            print(df)

        if df['KEY'].isnull().any() or df['TYP'].isnull().any():
            print("Missing value in'KEY' or 'TYP' column.")

        
        validation_row = pd.read_excel(file, engine='openpyxl', skiprows=5, nrows=1, usecols =lambda x: x not in ['Unnamed: 0', 'Unnamed: 1'])
        mandatory_columns = validation_row.columns[validation_row.eq('mandatory').any()]

        print(validation_row)

        print("Mandatory Columns:", mandatory_columns)

        

        df.reset_index(drop=True, inplace=True)
        missing_values = df[mandatory_columns].isnull()
        df_1 = df[~missing_values.any(axis=1)]

        #error_messages = []

        #for column in mandatory_columns:
            #if column not in ['KEY', 'TYP']:
                #missing_keys = df[df[column].isnull()]['KEY'].tolist()
                #if missing_keys:
                   # for key in missing_keys:
                    #    error_messages.append(f"Missing value in mandatory column '{column}' for KEY '{key}'")

        


       # if error_messages:
        #    raise ValueError('\n'.join(error_messages))


        print(df_1)

        json_obj = {}

        for i, row in df_1.iloc[1:].iterrows():
            row_data = {}
            for col in df.columns:
                if pd.notna(row[col]):
                    row_data[col] = row[col]
                json_obj[row['KEY']] = row_data
        json_str = json.dumps(json_obj, indent = 4)

        with open('output_datatemplate.txt', 'w') as file:
            file.write(data_str + '\n\n')
            file.write(json_str)
        
        print(json_str)


        missing_values = df[mandatory_columns].isnull()
        df_errors = df[missing_values.any(axis=1)]
        wb = load_workbook('Error_Template.xlsx')
        print(wb.sheetnames)
        
        
        ws = wb.active
        print(ws)
        df = pd.read_excel('Error_Template.xlsx', sheet_name=ws.title)
        print(df)
        
        current_row = 8

        for index, row in df_errors.iterrows():
            for i, value in enumerate(row):
                cell = ws.cell(row=current_row, column=i+2)
                if pd.isnull(value):
                    if ws.cell(row=6, column=i+2).value in ['KEY','TYP']:
                        cell.font = Font(color = "FF0000")
                        cell.value = "Critical Error"
                    else:
                        cell.value = "Mandatory Data"
                else:
                    cell.value = value
            current_row += 1




        wb.save(f'C:\\Users\\ASUS\\Desktop\\fieldmobi\\Error Folder\\Error_{timestamp}.xlsx')


except Exception as e:
    print(f"An error occured: {e}")





""" class MyHandler(FileSystemEventHandler):
    def on_modified(self, event):
        for filename in os.listdir(folder_to_track):
            src = folder_to_track + "/" + filename
            new_destination = folder_destination + "/" + filename
            os.rename(src, new_destination) """




""" 
folder_to_track = 'C:\\Users\\ASUS\\Desktop\\fieldmobi\\Upload Folder'
folder_destination = 'C:\\Users\\ASUS\\Desktop\\fieldmobi\\Error Folder'
event_handler = MyHandler()
observer = Observer()
observer.schedule(event_handler, folder_to_track, recursive=True)
observer.start()

try:
    while True:
        time.sleep(10)
except KeyboardInterrupt:
    observer.stop()
observer.join() """
               